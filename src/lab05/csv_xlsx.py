import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def csv_to_xlsx(csv_path: str, xlsx_path: str) -> None:
    """
    Конвертирует CSV в XLSX.
    Использовать openpyxl ИЛИ xlsxwriter.
    Первая строка CSV — заголовок.
    Лист называется "Sheet1".
    Колонки — автоширина по длине текста (не менее 8 символов).
    """
    cpath, xpath = Path(csv_path), Path(xlsx_path)
    if not cpath.exists():
        raise FileNotFoundError(cpath)
    if cpath.suffix != ".csv" or xpath.suffix != ".xlsx":
        raise ValueError

    with cpath.open(encoding="utf-8", newline="") as f:
        rows = list(csv.reader(f))

    if not rows:
        raise ValueError

    wb, ws = Workbook(), Workbook().active
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    for r in rows:
        ws.append(r)

    for i, col in enumerate(ws.columns, 1):
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[get_column_letter(i)].width = max(8, max_len)

    wb.save(xpath)
